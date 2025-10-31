from datetime import date, timedelta
import json

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib.auth.views import LoginView
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Max
from django.http import Http404, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from urllib.parse import urlparse
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.http import require_POST, require_http_methods

from .forms import DocumentoUploadForm, DocumentoForm
from .models import RegistroDocumento, Shopping, AnexoDocumento
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from .utils import notificar_usuario_documento
from django.views.generic import ListView
from .models import DocumentoObrigatorio
from django.contrib.auth.mixins import LoginRequiredMixin


@login_required
def painel_documentos(request):
    if not request.user.groups.filter(name='Corporativo').exists():
        return redirect('painel_redirect')

    hoje = timezone.now().date()
    # Alinhar com o Dashboard: usar janela de 60 dias (opcionalmente aceitar GET)
    try:
        janela = int(request.GET.get('janela', '60'))
    except (ValueError, TypeError):
        janela = 60

    dados_painel = []

    # Indicadores gerais
    total_geral = 0
    no_prazo_geral = 0
    vencidos_geral = 0
    a_vencer_geral = 0

    # Opcionalmente filtrar por status, como no Dashboard
    filtro_status = request.GET.get('status', '')

    for shopping in Shopping.objects.all():
        # Alinhar com Dashboard: considerar todos os registros (e permitir filtro de status)
        registros = RegistroDocumento.objects.filter(shopping=shopping)
        if filtro_status in ("pendente", "aprovado"):
            registros = registros.filter(status_aprovacao=filtro_status)

        total = registros.count()
        vencidos = registros.filter(data_vencimento__lt=hoje).count()
        a_vencer_qs = registros.filter(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=janela))
        a_vencer = a_vencer_qs.count()
        # Alinhar com Dashboard: no_prazo = total - (vencidos + a_vencer)
        no_prazo = total - (vencidos + a_vencer)

        # Status visual do painel
        if vencidos > 0:
            status = 'vermelho'
        elif a_vencer > 0:
            status = 'amarelo'
        else:
            status = 'verde'

        dados_painel.append({
            'shopping': shopping,
            'status': status,
            'total': total,
            'vencidos': vencidos,
            'a_vencer': a_vencer,
            'no_prazo': no_prazo,
        })

        total_geral += total
        no_prazo_geral += no_prazo
        vencidos_geral += vencidos
        a_vencer_geral += a_vencer

    return render(request, 'documentos/painel_documentos.html', {
        'dados_painel': dados_painel,
        'indicadores': {
            'total': total_geral,
            'no_prazo': no_prazo_geral,
            'a_vencer': a_vencer_geral,
            'vencidos': vencidos_geral,
        }
    })


@login_required
def painel_documentos_status_json(request):
    """Retorna um JSON com o status atual dos DocumentosObrigatorio visíveis ao usuário.
    Usado pelo painel para atualizar linhas dinamicamente via JS.
    """
    user = request.user
    qs = DocumentoObrigatorio.objects.filter(ativo=True)
    if user.groups.filter(name='Corporativo').exists() or user.is_superuser:
        qs = qs.order_by('area', 'categoria', 'nome')
    else:
        shopping = None
        if user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil'):
            shopping = user.perfil.shopping
        elif user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping'):
            shopping = user.shopping
        if shopping:
            qs = qs.filter(shopping=shopping).order_by('area', 'categoria', 'nome')
        else:
            qs = DocumentoObrigatorio.objects.none()

    resultado = []
    for doc in qs:
        # Se houver um documento avulso vinculado, usar ele como base de status/detalhes
        linked = getattr(doc, 'documento_vinculado', None)
        if linked:
            enviado = True
            ultimo = linked
        else:
            enviado = RegistroDocumento.objects.filter(shopping=doc.shopping, titulo=doc.nome).exists()
            ultimo = RegistroDocumento.objects.filter(shopping=doc.shopping, titulo=doc.nome).order_by('-data_envio').first()

        anexo = None
        detalhe_url = ''
        if ultimo:
            anexos = list(ultimo.anexos.all())
            if anexos:
                anexo = anexos[0]
            try:
                detalhe_url = reverse('detalhar_documento', args=[ultimo.tipo, ultimo.id])
            except Exception:
                detalhe_url = ''

        resultado.append({
            'id': doc.id,
            'nome': doc.nome,
            'area': doc.area,
            'categoria': doc.categoria,
            # status: 'concluido' quando existe registro, caso contrário 'pendente'
            'status': 'concluido' if enviado else 'pendente',
            'shopping_name': doc.shopping.nome,
            'anexo_url': anexo.arquivo.url if anexo else '',
            'detalhe_url': detalhe_url,
            'data_envio': ultimo.data_envio.strftime('%d/%m/%Y %H:%M') if ultimo and ultimo.data_envio else '',
            'ultima_atualizacao': (ultimo.data_aprovacao.strftime('%d/%m/%Y %H:%M') if ultimo and ultimo.data_aprovacao else (doc.marcado_em.strftime('%d/%m/%Y %H:%M') if doc.marcado_em else ''))
        })

    return JsonResponse(resultado, safe=False)


@login_required
@require_POST
def excluir_documento_obrigatorio(request, doc_obrig_id):
    """Exclui um DocumentoObrigatorio. Permissões: Corporativo, superuser, ou Gestor do mesmo shopping."""
    user = request.user
    doc = get_object_or_404(DocumentoObrigatorio, id=doc_obrig_id)

    permitted = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        permitted = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and user.perfil.shopping == doc.shopping:
        permitted = True

    if not permitted:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)
        messages.error(request, 'Permissão negada para excluir este documento.')
        return redirect('painel_documentos_gerencial')

    nome = doc.nome
    doc.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})

    messages.success(request, f'Documento obrigatório "{nome}" excluído com sucesso.')
    return redirect('painel_documentos_gerencial')


@login_required
def buscar_documentos_avulsos_json(request, doc_obrig_id):
    """Busca documentos avulsos do mesmo shopping para vinculação (JSON)."""
    doc_obrig = get_object_or_404(DocumentoObrigatorio, id=doc_obrig_id)
    q = request.GET.get('q', '').strip()
    qs = RegistroDocumento.objects.filter(shopping=doc_obrig.shopping)
    if q:
        qs = qs.filter(titulo__icontains=q)
    qs = qs.order_by('-data_envio')[:20]
    dados = []
    for d in qs:
        dados.append({
            'id': d.id,
            'titulo': d.titulo,
            'tipo': d.get_tipo_display() if hasattr(d, 'get_tipo_display') else d.tipo,
            'status': d.status_aprovacao,
            'data_envio': d.data_envio.strftime('%d/%m/%Y %H:%M') if d.data_envio else '',
        })
    return JsonResponse({'results': dados}, safe=True)


@login_required
@require_POST
def vincular_documento_obrigatorio(request, doc_obrig_id):
    """Vincula manualmente um RegistroDocumento avulso a uma premissa (DocumentoObrigatorio)."""
    user = request.user
    doc_obrig = get_object_or_404(DocumentoObrigatorio, id=doc_obrig_id)
    doc_id = request.POST.get('doc_id')

    # Permissões: Corporativo/superuser ou Gestor do mesmo shopping
    permitted = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        permitted = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and user.perfil.shopping == doc_obrig.shopping:
        permitted = True

    if not permitted:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    try:
        doc = RegistroDocumento.objects.get(id=doc_id, shopping=doc_obrig.shopping)
    except RegistroDocumento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Documento não encontrado neste shopping'}, status=404)

    doc_obrig.documento_vinculado = doc
    # Registrar data/autor do vínculo para aparecer como última atualização
    try:
        from django.utils import timezone
        doc_obrig.marcado_em = timezone.now()
        doc_obrig.marcado_por = user
    except Exception:
        pass
    doc_obrig.save()

    return JsonResponse({'success': True})


@login_required
@require_POST
def desvincular_documento_obrigatorio(request, doc_obrig_id):
    """Remove a vinculação manual de um RegistroDocumento avulso da premissa."""
    user = request.user
    doc_obrig = get_object_or_404(DocumentoObrigatorio, id=doc_obrig_id)

    permitted = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        permitted = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and user.perfil.shopping == doc_obrig.shopping:
        permitted = True

    if not permitted:
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    doc_obrig.documento_vinculado = None
    try:
        doc_obrig.marcado_em = None
        doc_obrig.marcado_por = None
    except Exception:
        pass
    doc_obrig.save()

    return JsonResponse({'success': True})

@login_required
def detalhes_shopping(request, shopping_id):
    shopping = get_object_or_404(Shopping, id=shopping_id)

    # Filtro por tipo
    tipo_filtro = request.GET.get('tipo', 'todos')

    registros = RegistroDocumento.objects.filter(shopping=shopping)
    if tipo_filtro in ['documento', 'contrato']:
        registros = registros.filter(tipo=tipo_filtro)

    # Obter somente os registros mais recentes por título
    ultimos_ids = (
        registros.values('titulo')
        .annotate(ultimo_id=Max('id'))
        .values_list('ultimo_id', flat=True)
    )

    registros_finais = RegistroDocumento.objects.filter(id__in=ultimos_ids).order_by('data_vencimento')

    # Paginação
    paginator = Paginator(registros_finais, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    hoje = date.today()
    hoje_plus_60 = hoje + timedelta(days=60)

    # determinar next para retornar quando necessário
    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER')

    return render(request, 'documentos/detalhes_shopping.html', {
        'shopping': shopping,
        'page_obj': page_obj,
        'tipo_filtro': tipo_filtro,
        'is_gestor': request.user.groups.filter(name='Gestor').exists(),
        'hoje': hoje,
        'hoje_plus_60': hoje_plus_60,
        'next': next_url,
    })

@login_required
@require_POST
def aprovar_documento(request, doc_id):
    user = request.user

    if not user.groups.filter(name='Gestor').exists():
        return redirect('painel_redirect')

    doc = get_object_or_404(RegistroDocumento, id=doc_id)

    if doc.shopping != user.perfil.shopping:
        messages.error(request, "Você não tem permissão para aprovar este documento.")
        return redirect('painel_redirect')

    if doc.status_aprovacao.strip().lower() != 'pendente':
        messages.error(request, "Apenas documentos pendentes podem ser aprovados.")
    else:
        doc.status_aprovacao = 'aprovado'
        doc.aprovado_por = user
        doc.data_aprovacao = timezone.localtime()
        doc.save()

        try:
            notificar_usuario_documento(doc, tipo='aprovado')
        except Exception as e:
            print(f"Erro ao enviar e-mail: {e}")

        messages.success(request, "✅ Documento aprovado com sucesso!")

    referer = request.META.get('HTTP_REFERER', '')
    if 'pendencias' in referer:
        return redirect('pendencias_gestor')
    return redirect('detalhes_shopping', shopping_id=doc.shopping.id)

@login_required
@require_http_methods(["GET", "POST"])
def novo_documento(request, shopping_id=None):
    user = request.user
    shopping = None

    if shopping_id:
        # Quando shopping_id é fornecido na URL
        shopping = get_object_or_404(Shopping, id=shopping_id)
        
        # Verificar se o usuário tem permissão para este shopping
        if not (user.is_superuser or 
                user.groups.filter(name='Corporativo').exists() or
                (user.groups.filter(name='Gestor').exists() and user.perfil.shopping == shopping) or
                (user.groups.filter(name='Usuario').exists() and user.shopping == shopping)):
            messages.error(request, "🚫 Você não tem permissão para adicionar documentos neste shopping.")
            return redirect('painel_redirect')
    else:
        # Lógica original quando shopping_id não é fornecido
        if user.groups.filter(name='Usuario').exists():
            shopping = user.shopping
        elif user.groups.filter(name='Gestor').exists():
            shopping = user.perfil.shopping
        elif user.groups.filter(name='Corporativo').exists() or user.is_superuser:
            # Para corporativo/superuser, precisa escolher o shopping
            shopping_id = request.GET.get('shopping_id')
            if shopping_id:
                shopping = get_object_or_404(Shopping, id=shopping_id)
            else:
                # Redireciona para escolher o shopping
                messages.info(request, "📍 Selecione um shopping para adicionar o documento.")
                return redirect('painel_geral')
        else:
            return redirect('painel_redirect')

    # gerenciar retorno dinâmico
    next_url = request.POST.get('next') if request.method == 'POST' else request.GET.get('next') or request.META.get('HTTP_REFERER')

    if request.method == 'POST':
        form = DocumentoUploadForm(request.POST)
        anexos = request.FILES.getlist('anexos')
        
        if form.is_valid() and anexos:
            dados = form.cleaned_data

            novo_doc = RegistroDocumento.objects.create(
                shopping=shopping,
                titulo=dados['titulo'],
                tipo=dados['tipo'],
                descricao=dados['descricao'],
                data_emissao=dados['data_emissao'],
                data_vencimento=dados['data_vencimento'],
                enviado_por=user
            )

            # Salva os anexos
            for arquivo in anexos:
                AnexoDocumento.objects.create(documento=novo_doc, arquivo=arquivo)

            try:
                grupo_gestor = Group.objects.get(name="Gestor")
                gestores = grupo_gestor.user_set.filter(perfil__shopping=shopping)
            except Group.DoesNotExist:
                gestores = []

            destinatarios = [g.email for g in gestores if g.email]

            if destinatarios:
                assunto = f"\U0001F4E5 Novo {dados['tipo']} enviado - {dados['titulo']}"
                mensagem = (
                    f"Olá,\n\n"
                    f"Um novo {dados['tipo']} foi enviado por {user.get_full_name() or user.username}.\n\n"
                    f"Título: {dados['titulo']}\n"
                    f"Data de Vencimento: {dados['data_vencimento']}\n"
                    f"Anexos: {len(anexos)} arquivo(s)\n\n"
                    f"Acesse o painel para aprovar:\n"
                    f"{request.build_absolute_uri(reverse('detalhes_shopping', args=[shopping.id]))}"
                )
                send_mail(
                    assunto,
                    mensagem,
                    'nao-responda@sistema.com',
                    destinatarios,
                    fail_silently=True
                )

            messages.success(request, f"🎉 Documento enviado com sucesso! {len(anexos)} anexo(s) adicionado(s).")
            # redirecionar para next quando seguro
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                return redirect(next_url)
            return redirect('detalhes_shopping', shopping_id=shopping.id)
        else:
            if not anexos:
                messages.error(request, "⚠️ É necessário enviar pelo menos um arquivo.")
    else:
        # Permite pre-preencher o formulário via querystring, por exemplo ?titulo=Nome
        initial = {}
        titulo_prefill = request.GET.get('titulo') or request.GET.get('titulo')
        tipo_prefill = request.GET.get('tipo')
        if titulo_prefill:
            initial['titulo'] = titulo_prefill
        if tipo_prefill:
            initial['tipo'] = tipo_prefill
        form = DocumentoUploadForm(initial=initial)

    return render(request, 'documentos/novo_documento.html', {'form': form})

@login_required
def pendencias_gestor(request):
    user = request.user
    if not user.groups.filter(name='Gestor').exists():
        return redirect('painel_redirect')

    shopping = user.perfil.shopping
    hoje = timezone.now().date()

    # Buscar apenas documentos pendentes
    pendentes = RegistroDocumento.objects.filter(shopping=shopping, status_aprovacao='pendente').order_by('-data_envio')

    # Paginação
    paginator = Paginator(pendentes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Indicadores
    todos = RegistroDocumento.objects.filter(shopping=shopping)
    vencidos = todos.filter(data_vencimento__lt=hoje).count()
    a_vencer = todos.filter(data_vencimento__range=[hoje, hoje + timedelta(days=60)]).count()
    no_prazo = todos.filter(data_vencimento__gt=hoje + timedelta(days=60)).count()
    aprovados = todos.filter(status_aprovacao='aprovado').count()

    return render(request, 'documentos/pendencias_gestor.html', {
        'pendencias': page_obj,
        'page_obj': page_obj,
        'is_gestor': True,
        'indicadores': {
            'total': todos.count(),
            'vencidos': vencidos,
            'a_vencer': a_vencer,
            'no_prazo': no_prazo,
            'aprovados': aprovados,
        }
    })

@login_required
def painel_redirect(request):
    user = request.user

    if user.groups.filter(name='Corporativo').exists():
        return redirect('painel_geral')

    elif user.groups.filter(name='Gestor').exists():
        if hasattr(user, 'perfil') and user.perfil.shopping:
            return redirect('pendencias_gestor')
        else:
            messages.error(request, "Seu perfil de gestor não está configurado corretamente.")
            logout(request)
            return redirect('login')

    elif user.groups.filter(name='Usuario').exists():
        if hasattr(user, 'shopping') and user.shopping:
            return redirect('detalhes_shopping', shopping_id=user.shopping.id)
        else:
            messages.error(request, "Usuário comum sem shopping associado.")
            logout(request)
            return redirect('login')

    else:
        messages.error(request, "Seu usuário não possui grupo válido.")
        logout(request)
        return redirect('login')
    

class CustomLoginView(auth_views.LoginView):
    template_name = 'registration/login.html'

    def form_valid(self, form):
        messages.success(self.request, "Login realizado com sucesso!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Usuário ou senha inválidos. Tente novamente.")
        return super().form_invalid(form)

class CustomLogoutView(auth_views.LogoutView):
    def dispatch(self, request, *args, **kwargs):
        messages.success(request, "Logout realizado com sucesso!")
        return super().dispatch(request, *args, **kwargs)

    
@login_required
@require_http_methods(["GET", "POST"])
def reprovar_documento(request, doc_id):
    user = request.user

    if not user.groups.filter(name='Gestor').exists():
        return redirect('painel_redirect')

    doc = get_object_or_404(RegistroDocumento, id=doc_id)

    if doc.shopping != user.perfil.shopping:
        messages.error(request, "Você não tem permissão para reprovar este documento.")
        return redirect('painel_redirect')

    if request.method == 'POST':
        if doc.status_aprovacao.strip().lower() != 'pendente':
            messages.error(request, "Apenas documentos pendentes podem ser reprovados.")
            referer = request.META.get('HTTP_REFERER', '')
            return redirect('pendencias_gestor' if 'pendencias' in referer else 'detalhes_shopping', shopping_id=doc.shopping.id)

        motivo = request.POST.get('motivo', '').strip()
        doc.status_aprovacao = 'reprovado'
        doc.motivo_reprovacao = motivo
        doc.aprovado_por = user
        doc.data_aprovacao = timezone.localtime()
        doc.save()

        try:
            notificar_usuario_documento(doc, tipo='reprovado')
        except Exception as e:
            print(f"Erro ao enviar e-mail: {e}")

        messages.warning(request, "Documento reprovado com sucesso.")
        return redirect('pendencias_gestor')

    return render(request, 'documentos/form_reprovar.html', {'doc': doc})

@login_required
def editar_documento(request, tipo, doc_id):
    # Buscar o documento sem restringir ao autor; aplicar checagem de permissão abaixo
    doc = get_object_or_404(RegistroDocumento, id=doc_id)
    next_url = request.POST.get('next') if request.method == 'POST' else request.GET.get('next') or request.META.get('HTTP_REFERER')

    # Permissões: Superuser ou Corporativo sempre podem editar; Gestor do mesmo shopping; autor pode editar
    pode_editar = False
    user = request.user
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        pode_editar = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and getattr(user.perfil, 'shopping', None) and doc.shopping_id == user.perfil.shopping_id:
        pode_editar = True
    elif doc.enviado_por_id == user.id:
        pode_editar = True

    if not pode_editar:
        messages.error(request, 'Você não tem permissão para editar este documento.')
        # Redireciona para detalhar o documento, preservando next quando seguro
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
            return redirect(next_url)
        return redirect('detalhar_documento', tipo=doc.tipo, doc_id=doc.id)

    if request.method == 'POST':
        form = DocumentoForm(request.POST, instance=doc)
        novos_anexos = request.FILES.getlist('anexos')
        
        if form.is_valid():
            form.save()
            
            # Processar novos anexos se houver
            if novos_anexos:
                for arquivo in novos_anexos:
                    AnexoDocumento.objects.create(documento=doc, arquivo=arquivo)
                messages.success(request, f'✅ Documento atualizado com sucesso! {len(novos_anexos)} novo(s) anexo(s) adicionado(s).')
            else:
                messages.success(request, '✅ Documento atualizado com sucesso!')
            
            # redirecionar para next quando seguro
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                return redirect(next_url)
            return redirect('detalhar_documento', tipo=doc.tipo, doc_id=doc.id)

    else:
        form = DocumentoForm(instance=doc)

    return render(request, 'documentos/editar_documento.html', {
        'form': form,
        'doc': doc,
        'hoje': date.today(),
        'hoje_plus_60': date.today() + timedelta(days=60),
        'tipo_objeto': tipo,  # necessário para o template
    })

@login_required
@require_POST
def excluir_documento(request, tipo, doc_id):
    doc = get_object_or_404(RegistroDocumento, id=doc_id)

    # Permissões: Superuser, Corporativo, Gestor do mesmo shopping, ou autor do documento
    user = request.user
    pode_excluir = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        pode_excluir = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and getattr(user.perfil, 'shopping', None) and doc.shopping_id == user.perfil.shopping_id:
        pode_excluir = True
    elif doc.enviado_por_id == user.id:
        pode_excluir = True

    if pode_excluir:
        doc.delete()
        messages.success(request, "🗑️ Documento excluído com sucesso.")
    else:
        messages.error(request, "🚫 Você não tem permissão para excluir este documento.")

    return redirect('detalhes_shopping', shopping_id=doc.shopping.id)

@login_required
@require_POST
def excluir_anexo(request, anexo_id):
    """View para excluir anexo individual via AJAX"""
    from django.http import JsonResponse
    
    try:
        anexo = get_object_or_404(AnexoDocumento, id=anexo_id)
        documento = anexo.documento
        
        # Verificar permissão: Superuser, Corporativo, Gestor do mesmo shopping, ou autor
        user = request.user
        pode_excluir_anexo = False
        if user.is_superuser or user.groups.filter(name='Corporativo').exists():
            pode_excluir_anexo = True
        elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and getattr(user.perfil, 'shopping', None) and documento.shopping_id == user.perfil.shopping_id:
            pode_excluir_anexo = True
        elif documento.enviado_por_id == user.id:
            pode_excluir_anexo = True

        if not pode_excluir_anexo:
            return JsonResponse({'success': False, 'error': '🚫 Você não tem permissão para excluir este anexo'})
        
        # Excluir o arquivo físico se existir
        if anexo.arquivo and anexo.arquivo.name:
            try:
                anexo.arquivo.delete(save=False)
            except Exception:
                pass  # Continua mesmo se não conseguir excluir o arquivo físico
        
        # Excluir o registro do anexo
        anexo.delete()
        
        return JsonResponse({'success': True, 'message': '✅ Anexo excluído com sucesso!'})
        
    except AnexoDocumento.DoesNotExist:
        return JsonResponse({'success': False, 'error': '❌ Anexo não encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'💥 Erro interno: {str(e)}'})

@login_required
def confirmar_exclusao(request, tipo, doc_id):
    doc = get_object_or_404(RegistroDocumento, id=doc_id)
    # Permissões: Superuser, Corporativo, Gestor do mesmo shopping, ou autor do documento
    user = request.user
    pode_excluir = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        pode_excluir = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and getattr(user.perfil, 'shopping', None) and doc.shopping_id == user.perfil.shopping_id:
        pode_excluir = True
    elif doc.enviado_por_id == user.id:
        pode_excluir = True

    if not pode_excluir:
        messages.error(request, "🚫 Você não tem permissão para excluir este documento.")
        return redirect('detalhar_documento', tipo=tipo, doc_id=doc_id)

    if request.method == 'POST':
        shopping_id = doc.shopping.id
        doc.delete()
        messages.success(request, "🗑️ Documento excluído com sucesso!")
        return redirect('detalhes_shopping', shopping_id=shopping_id)

    return render(request, 'documentos/confirmar_exclusao.html', {'doc': doc})

@login_required
def detalhar_documento(request, tipo, doc_id):
    doc = get_object_or_404(RegistroDocumento, id=doc_id)

    hoje = date.today()
    hoje_plus_60 = hoje + timedelta(days=60)

    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER')
    # Calcular URL segura de retorno para evitar loops entre editar/detalhar
    current_path = request.path
    back_url = None
    if next_url:
        try:
            parsed = urlparse(next_url)
            next_path = parsed.path or next_url
        except Exception:
            next_path = next_url
        edit_path = reverse('editar_documento', kwargs={'tipo': tipo, 'doc_id': doc_id})
        if next_path == current_path or next_path == edit_path:
            back_url = reverse('detalhes_shopping', args=[doc.shopping.id])
        else:
            back_url = next_url
    else:
        back_url = reverse('detalhes_shopping', args=[doc.shopping.id])
    # Flags de permissão para o template
    user = request.user
    pode_editar = False
    if user.is_superuser or user.groups.filter(name='Corporativo').exists():
        pode_editar = True
    elif user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and getattr(user.perfil, 'shopping', None) and doc.shopping_id == user.perfil.shopping_id:
        pode_editar = True
    elif doc.enviado_por_id == user.id:
        pode_editar = True

    pode_excluir = pode_editar  # Regra: quem pode editar também pode excluir
    return render(request, 'documentos/detalhar_documento.html', {
        'doc': doc,
        'hoje': hoje,
        'hoje_plus_60': hoje_plus_60,
        'tipo_objeto': tipo,  # para o template renderizar o tipo corretamente
        'next': next_url,
        'pode_editar': pode_editar,
        'pode_excluir': pode_excluir,
        'back_url': back_url,
    })

@login_required
def reenviar_para_aprovacao(request, doc_id):
    doc = get_object_or_404(RegistroDocumento, id=doc_id)

    if doc.status_aprovacao == 'reprovado':
        doc.status_aprovacao = 'pendente'
        doc.motivo_reprovacao = ''
        doc.data_envio = timezone.now()
        doc.aprovado_por = None
        doc.data_aprovacao = None
        doc.save()
        messages.success(request, '📤 Documento reenviado para aprovação com sucesso.')
    else:
        messages.warning(request, '❌ Apenas documentos reprovados podem ser reenviados.')

    return redirect('detalhar_documento', tipo=doc.tipo, doc_id=doc.id)

def ver_historico_documento(request, titulo, shopping_id):
    shopping = Shopping.objects.get(id=shopping_id)
    documentos = RegistroDocumento.objects.filter(titulo=titulo, shopping=shopping).order_by('-data_envio')

    # Fallback: se não houver histórico por título, tentar exibir o documento
    # avulso vinculado à premissa correspondente para que o usuário veja algo
    # útil ao clicar em "Histórico" a partir do painel gerencial.
    if not documentos.exists():
        try:
            doc_obrig = DocumentoObrigatorio.objects.filter(shopping=shopping, nome=titulo).first()
        except Exception:
            doc_obrig = None
        if doc_obrig and doc_obrig.documento_vinculado:
            # Use o documento vinculado como item único de histórico
            documentos = [doc_obrig.documento_vinculado]

    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER')
    return render(request, 'documentos/historico_documento.html', {
        'documentos': documentos,
        'titulo': titulo,
        'tipo': documentos[0].tipo if documentos else '',
        'doc_id': documentos[0].id if documentos else '',
        'shopping_id': shopping_id,
        'next': next_url,
    })


@login_required
@require_POST
def marcar_documento_obrigatorio(request, doc_obrig_id):
    """Alterna a marcação manual (ok) feita pelo gestor para um DocumentoObrigatorio."""
    user = request.user
    # Somente Gestor ou Corporativo pode marcar
    if not (user.groups.filter(name='Gestor').exists() or user.groups.filter(name='Corporativo').exists() or user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Permissão negada'}, status=403)

    doc = get_object_or_404(DocumentoObrigatorio, id=doc_obrig_id)
    # Alterna
    doc.marcado_ok = not doc.marcado_ok
    if doc.marcado_ok:
        doc.marcado_por = user
        doc.marcado_em = timezone.now()
    else:
        doc.marcado_por = None
        doc.marcado_em = None
    doc.save()

    # Se requisição AJAX, retorna JSON; caso contrário, redireciona para o painel com mensagem
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'marcado_ok': doc.marcado_ok})

    # Requisição padrão (form POST) — usar messages e redirect
    if doc.marcado_ok:
        messages.success(request, f'Documento "{doc.nome}" marcado como OK.')
    else:
        messages.info(request, f'Documento "{doc.nome}" desmarcado.')

    # Redireciona de volta ao painel gerencial
    return redirect('painel_documentos_gerencial')


class PainelDocumentosView(LoginRequiredMixin, ListView):
    """Painel visual para o gestor acompanhar documentos obrigatórios (somente leitura)."""
    model = DocumentoObrigatorio
    template_name = 'documentos/painel_documentos_gerencial.html'
    context_object_name = 'documentos'

    def get_queryset(self):
        user = self.request.user
        # Gestor vê apenas seu shopping; Corporativo vê todos
        if user.groups.filter(name='Corporativo').exists() or user.is_superuser:
            return DocumentoObrigatorio.objects.filter(ativo=True).order_by('area', 'categoria', 'nome')

        # Para Gestor e Usuario, tentar inferir shopping via perfil/shopping
        shopping = None
        if user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil'):
            shopping = user.perfil.shopping
        elif user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping'):
            shopping = user.shopping

        if shopping:
            return DocumentoObrigatorio.objects.filter(ativo=True, shopping=shopping).order_by('area', 'categoria', 'nome')

        return DocumentoObrigatorio.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        request = self.request
        filtro_area = request.GET.get('area', '').strip()
        filtro_categoria = request.GET.get('categoria', '').strip()
        filtro_shopping = request.GET.get('shopping', '').strip()
        filtro_status = request.GET.get('status', '').strip().lower()

        # base queryset já considera o escopo do usuário (get_queryset())
        base_qs = self.get_queryset()

        # aplicar filtros no banco quando possível
        if filtro_area:
            base_qs = base_qs.filter(area__icontains=filtro_area)
        if filtro_categoria:
            base_qs = base_qs.filter(categoria__icontains=filtro_categoria)
        if filtro_shopping:
            try:
                base_qs = base_qs.filter(shopping__id=int(filtro_shopping))
            except Exception:
                pass

        documentos = []
        total_count = 0
        concluidos_count = 0
        pendentes_count = 0
        for doc in base_qs:
            # Preferir documento vinculado, quando existir
            linked = getattr(doc, 'documento_vinculado', None)
            if linked:
                enviado = True
                ultimo = linked
            else:
                enviado = RegistroDocumento.objects.filter(shopping=doc.shopping, titulo=doc.nome).exists()
                ultimo = RegistroDocumento.objects.filter(shopping=doc.shopping, titulo=doc.nome).order_by('-data_envio').first()
            anexos = ultimo.anexos.all() if ultimo else []
            # Considerar como concluído se houver envio/vínculo OU marcado manualmente como OK
            status = 'concluido' if (enviado or doc.marcado_ok) else 'pendente'

            # KPIs agregados
            total_count += 1
            if status == 'concluido':
                concluidos_count += 1
            else:
                pendentes_count += 1

            # aplicar filtro de status em memória
            if filtro_status:
                if filtro_status == 'pendente' and status != 'pendente':
                    continue
                if filtro_status == 'concluido' and status != 'concluido':
                    continue

            documentos.append({
                'obj': doc,
                'status': status,
                'anexo': anexos[0] if anexos else None,
                'data_envio': ultimo.data_envio if ultimo else None,
                'ultima_atualizacao': ultimo.data_aprovacao if ultimo else (doc.marcado_em if doc.marcado_em else None),
                'historico_titulo': (ultimo.titulo if ultimo else doc.nome),
                'historico_shopping_id': doc.shopping.id,
            })

        # listas para filtros
        all_qs = DocumentoObrigatorio.objects.filter(ativo=True)
        ctx['area_list'] = sorted(set([a for a in all_qs.values_list('area', flat=True) if a]))
        ctx['categoria_list'] = sorted(set([c for c in all_qs.values_list('categoria', flat=True) if c]))

        # lista de shoppings para o formulário rápido: todos para corporativo, ou apenas o do gestor
        if self.request.user.groups.filter(name='Corporativo').exists() or self.request.user.is_superuser:
            ctx['shopping_list'] = Shopping.objects.all()
        else:
            sh = None
            if self.request.user.groups.filter(name='Gestor').exists() and hasattr(self.request.user, 'perfil'):
                sh = self.request.user.perfil.shopping
            elif self.request.user.groups.filter(name='Usuario').exists() and hasattr(self.request.user, 'shopping'):
                sh = self.request.user.shopping
            ctx['shopping_list'] = [sh] if sh else []

        # Paginação
        page_number = request.GET.get('page')
        per_page = getattr(self, 'paginate_by', 10) or 10
        paginator = Paginator(documentos, per_page)
        page_obj = paginator.get_page(page_number)

        ctx['page_obj'] = page_obj
        ctx['paginator'] = paginator
        ctx['is_paginated'] = page_obj.has_other_pages()
        ctx['documentos'] = page_obj.object_list

        # manter filtros no contexto
        ctx['filtro_area'] = filtro_area
        ctx['filtro_categoria'] = filtro_categoria
        ctx['filtro_shopping'] = filtro_shopping
        ctx['filtro_status'] = filtro_status

        # Querystring sem parâmetro de página para evitar duplicação de "page" nos links de paginação
        try:
            params = request.GET.copy()
            if 'page' in params:
                params.pop('page')
            ctx['qs_no_page'] = params.urlencode()
        except Exception:
            ctx['qs_no_page'] = ''

        # KPIs do painel (com base no conjunto filtrado)
        ctx['kpis'] = {
            'total': total_count,
            'pendentes': pendentes_count,
            'concluidos': concluidos_count,
            'percent_concluidos': round((concluidos_count / total_count) * 100, 1) if total_count else 0.0,
            'percent_pendentes': round((pendentes_count / total_count) * 100, 1) if total_count else 0.0,
        }

        return ctx

from django import forms
from django.utils.http import url_has_allowed_host_and_scheme
import io
import openpyxl

class UploadExcelForm(forms.Form):
    arquivo = forms.FileField(label='Arquivo Excel (.xlsx)')
    # opcional: permite escolher shopping quando corporativo
    shopping = forms.ModelChoiceField(queryset=Shopping.objects.all(), required=False)

    def __init__(self, *args, user=None, **kwargs):
        super().__init__(*args, **kwargs)
        # Se for Gestor ou Usuario, limitar a lista de shoppings ao do usuário
        try:
            if user is not None:
                if user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and user.perfil.shopping:
                    self.fields['shopping'].queryset = Shopping.objects.filter(id=user.perfil.shopping.id)
                    self.fields['shopping'].initial = user.perfil.shopping
                elif user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping') and user.shopping:
                    self.fields['shopping'].queryset = Shopping.objects.filter(id=user.shopping.id)
                    self.fields['shopping'].initial = user.shopping
                else:
                    # corporativo / superuser: manter todos
                    self.fields['shopping'].queryset = Shopping.objects.all()
        except Exception:
            # Em caso de qualquer erro, deixar o comportamento padrão
            self.fields['shopping'].queryset = Shopping.objects.all()


@login_required
@require_http_methods(["GET", "POST"])
def importar_documentos_excel(request):
    user = request.user
    # permitir redirecionamento de volta para a página anterior via parâmetro 'next'
    next_url = request.POST.get('next') if request.method == 'POST' else request.GET.get('next') or request.META.get('HTTP_REFERER')

    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            arquivo = request.FILES['arquivo']
            shopping = form.cleaned_data.get('shopping')
            # determinar shopping quando não informado
            if not shopping:
                if user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil'):
                    shopping = user.perfil.shopping
                elif user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping'):
                    shopping = user.shopping
                else:
                    # corporativo sem shopping escolhido
                    messages.error(request, 'Por favor selecione um shopping ao importar (usuários corporativos).')
                    # redirecionar de volta para next se for seguro
                    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                        return redirect(next_url)
                    return redirect('painel_documentos_gerencial')

            # processar excel com auto-mapeamento de colunas
            wb = openpyxl.load_workbook(filename=io.BytesIO(arquivo.read()), data_only=True)
            ws = wb.active
            # detectar cabeçalho na primeira linha
            header = [str(cell).strip().lower() if cell is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            # mapeamento de colunas esperadas
            def find_col(names):
                for n in names:
                    if n in header:
                        return header.index(n)
                return None

            idx_area = find_col(['área', 'area', 'area '])
            idx_categoria = find_col(['categoria', 'category'])
            idx_nome = find_col(['nome', 'documento', 'name', 'titulo'])
            idx_tipo = find_col(['tipo', 'type'])
            idx_shopping = find_col(['shopping', 'loja', 'mall', 'centro'])

            created = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # extrair valores com base nos índices detectados
                area = (row[idx_area] if idx_area is not None and idx_area < len(row) else None)
                categoria = (row[idx_categoria] if idx_categoria is not None and idx_categoria < len(row) else None)
                nome = (row[idx_nome] if idx_nome is not None and idx_nome < len(row) else None)
                tipo = (row[idx_tipo] if idx_tipo is not None and idx_tipo < len(row) else None)
                shopping_cell = (row[idx_shopping] if idx_shopping is not None and idx_shopping < len(row) else None)

                if not nome:
                    skipped += 1
                    continue

                # determinar shopping para esta linha: prioridade - coluna shopping na planilha, campo do form, inferência pelo usuário
                linha_shopping = None
                if shopping_cell:
                    # procurar shopping pelo nome (case-insensitive) ou sigla
                    sname = str(shopping_cell).strip()
                    try:
                        linha_shopping = Shopping.objects.filter(nome__iexact=sname).first()
                        if not linha_shopping:
                            linha_shopping = Shopping.objects.filter(sigla__iexact=sname).first()
                    except Exception:
                        linha_shopping = None

                if not linha_shopping:
                    linha_shopping = shopping

                if not linha_shopping:
                    # sem shopping para associar -> pular esta linha
                    skipped += 1
                    continue

                obj, created_flag = DocumentoObrigatorio.objects.get_or_create(
                    shopping=linha_shopping,
                    nome=str(nome).strip(),
                    defaults={'area': str(area).strip() if area else '', 'categoria': str(categoria).strip() if categoria else '', 'ativo': True}
                )
                if created_flag:
                    created += 1

            messages.success(request, f'Importação concluída. {created} novo(s) documento(s) criado(s). {skipped} linha(s) ignorada(s).')
            # redirecionar para next quando seguro, senão para o painel gerencial
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
                return redirect(next_url)
            return redirect('painel_documentos_gerencial')
    else:
        form = UploadExcelForm(user=request.user)

    return render(request, 'documentos/importar_documentos.html', {'form': form, 'next': next_url})


@login_required
@require_POST
def create_documento_obrigatorio(request):
    """Cria um DocumentoObrigatorio individualmente (usado pelo painel)."""
    user = request.user
    area = request.POST.get('area', '').strip()
    categoria = request.POST.get('categoria', '').strip()
    nome = request.POST.get('nome', '').strip()
    shopping_id = request.POST.get('shopping')
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER')

    if not nome:
        messages.error(request, 'Nome é obrigatório.')
        return redirect('painel_documentos_gerencial')

    # determinar shopping
    shopping = None
    if shopping_id:
        try:
            shopping = Shopping.objects.get(id=shopping_id)
        except Shopping.DoesNotExist:
            shopping = None

    if not shopping:
        # tentar inferir pelo perfil do usuário
        if user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil'):
            shopping = user.perfil.shopping
        elif user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping'):
            shopping = user.shopping

    if not shopping:
        messages.error(request, 'Selecione um shopping válido.')
        return redirect('painel_documentos_gerencial')

    # permissões: Corporativo, superuser, Gestor do shopping ou Usuario associado
    if not (user.is_superuser or
            user.groups.filter(name='Corporativo').exists() or
            (user.groups.filter(name='Gestor').exists() and hasattr(user, 'perfil') and user.perfil.shopping == shopping) or
            (user.groups.filter(name='Usuario').exists() and hasattr(user, 'shopping') and user.shopping == shopping)):
        messages.error(request, 'Permissão negada.')
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
            return redirect(next_url)
        return redirect('painel_documentos_gerencial')

    obj, created_flag = DocumentoObrigatorio.objects.get_or_create(
        shopping=shopping,
        nome=nome,
        defaults={'area': area or '', 'categoria': categoria or '', 'ativo': True}
    )

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'created': created_flag, 'id': obj.id, 'shopping_name': shopping.nome})

    if created_flag:
        messages.success(request, f'Documento obrigatório "{nome}" criado para {shopping.nome}.')
    else:
        messages.info(request, f'Documento obrigatório já existia para {shopping.nome}.')

    # usar next quando informado e seguro
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        return redirect(next_url)
    return redirect('painel_documentos_gerencial')

@login_required
def agenda_documentos(request):
    documentos = RegistroDocumento.objects.all()

    eventos = []
    for doc in documentos:
        dias_restantes = (doc.data_vencimento - date.today()).days

        # Cores por status:
        if dias_restantes < 0:
            cor = "#dc3545"  # vermelho - vencido
        elif dias_restantes <= 5:
            cor = "#ffc107"  # amarelo - próximo do vencimento
        else:
            cor = "#28a745"  # verde - tranquilo

        eventos.append({
            "title": doc.titulo,
            "start": doc.data_vencimento.strftime('%Y-%m-%d'),
            "color": cor,
            "url": reverse('detalhar_documento', kwargs={"tipo": doc.tipo, "doc_id": doc.id}),
        })

    return render(request, 'documentos/agenda.html', {
        "eventos_json": json.dumps(eventos),
    })


@login_required
def api_vencimentos_json(request):
    hoje = timezone.now().date()

    if request.user.groups.filter(name='Corporativo').exists():
        documentos = RegistroDocumento.objects.all()
    else:
        documentos = RegistroDocumento.objects.filter(shopping=request.user.perfil.shopping)

    eventos = []
    for doc in documentos:
        # ✅ Cores baseadas na data
        if doc.data_vencimento < hoje:
            cor = '#dc3545'  # vermelho
        elif doc.data_vencimento <= hoje + timedelta(days=60):
            cor = '#ffc107'  # amarelo
        else:
            cor = '#28a745'  # verde

        eventos.append({
            "title": doc.titulo,
            "start": doc.data_vencimento.isoformat(),
            "color": cor,
            "url": reverse('detalhar_documento', kwargs={
                'tipo': doc.tipo.id if hasattr(doc.tipo, 'id') else doc.tipo,
                'doc_id': doc.id
            })
        })

    return JsonResponse(eventos, safe=False)
